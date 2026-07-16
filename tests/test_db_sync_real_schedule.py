"""Tests for db/sync_real_schedule.py, using a small synthetic Master
Schedule workbook + roster CSV built with openpyxl.Workbook() — same
test-fixture convention as tests/test_real_schedule_assist_list.py (the one
place write-mode openpyxl is acceptable, per CLAUDE.md's real_schedule/
policy). Runs entirely against sqlite:///:memory:, never Resident_Schedules/.
"""

from __future__ import annotations

import datetime as dt

import openpyxl
import pytest

from db.models import Assignment, Block, Resident, Rotation, TimeOff, get_engine, get_session, init_db
from db.sync_real_schedule import sync

_HEADER_ROW = 3
_DATA_START_ROW = 4
_FIRST_WEEK_COL = 12

# AY2026-27 weeks (Mondays), plus one AY2025-26 week that must be excluded.
_PRIOR_YEAR_WEEK = dt.date(2026, 6, 29)
_WEEK_1 = dt.date(2026, 7, 6)
_WEEK_2 = dt.date(2026, 7, 13)
_WEEK_3 = dt.date(2026, 7, 20)
_WEEK_4 = dt.date(2026, 7, 27)
_ALL_WEEKS = [_PRIOR_YEAR_WEEK, _WEEK_1, _WEEK_2, _WEEK_3, _WEEK_4]


def _write_master_schedule(ws, rows: list[tuple[str, str, list[str | None]]]) -> None:
    """`rows` is a list of (name, pgy_label, [rotation_per_week...]) — one
    rotation value per entry in _ALL_WEEKS, aligned positionally."""
    for col_offset, week in enumerate(_ALL_WEEKS):
        ws.cell(row=_HEADER_ROW, column=_FIRST_WEEK_COL + col_offset, value=dt.datetime(week.year, week.month, week.day))
    row_num = _DATA_START_ROW
    for name, pgy_label, rotations in rows:
        ws.cell(row=row_num, column=1, value=name)
        ws.cell(row=row_num, column=5, value=pgy_label)
        for col_offset, rotation in enumerate(rotations):
            if rotation is not None:
                ws.cell(row=row_num, column=_FIRST_WEEK_COL + col_offset, value=rotation)
        row_num += 1


def _build_master_schedule_fixture(tmp_path, rows: list[tuple[str, str, list[str | None]]]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Intern Master")
    _write_master_schedule(ws, rows)
    # Upper Level Master must exist (load_master_schedule reads both
    # sheets) — a header row with no data rows is enough, matching how an
    # empty-but-present sheet looks in the real workbook.
    upper_ws = wb.create_sheet("Upper Level Master")
    _write_master_schedule(upper_ws, [])
    path = tmp_path / "master_MASTER_Schedule.xlsx"
    wb.save(path)
    return str(path)


def _build_roster_fixture(tmp_path, names: list[str]):
    path = tmp_path / "duke_residency.csv"
    path.write_text("Name,Phone,Email,Organization,Program,StartYear\n" + "".join(f"{n},,,,,\n" for n in names))
    return str(path)


@pytest.fixture
def session():
    engine = get_engine("sqlite:///:memory:")
    init_db(engine)
    with get_session(engine) as s:
        yield s


def test_sync_populates_residents_rotations_blocks_assignments(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [
            ("Chen, Alice", "PGY1", [None, "VA GM", "VA GM", "AMB Endo", "AMB Endo"]),
            ("Osei, Brian", "PGY1", [None, "VA GM", "VA GM", "VA GM", "VA GM"]),
        ],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen", "Brian Osei"])

    summary = sync(session, master_schedule_path=schedule_path, roster_path=roster_path)

    assert summary["residents"] == 2
    assert summary["blocks"] == 4  # the prior-AY week is excluded
    assert session.query(Resident).count() == 2
    assert {r.name for r in session.query(Resident).all()} == {"Chen, Alice", "Osei, Brian"}

    alice = session.query(Resident).filter_by(name="Chen, Alice").one()
    assert alice.pgy == 1

    # 4 real weeks each for 2 residents = 8 assignments (none of the 4 weeks
    # is a non-committing label here).
    assert session.query(Assignment).count() == 8

    va_gm = session.query(Rotation).filter_by(name="VA GM").one()
    # Week of 7/6 and 7/13: both residents on VA GM -> observed capacity 2.
    assert va_gm.intern_capacity == 2

    week1_block = session.query(Block).filter_by(year=2026, start_date=_WEEK_1).one()
    alice_assignment = (
        session.query(Assignment).filter_by(resident_id=alice.id, block_id=week1_block.id).one()
    )
    assert alice_assignment.rotation_id == va_gm.id
    assert alice_assignment.role == "intern"


def test_sync_excludes_prior_academic_year_week(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [("Chen, Alice", "PGY1", ["Neuro", "VA GM", "VA GM", "VA GM", "VA GM"])],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen"])

    sync(session, master_schedule_path=schedule_path, roster_path=roster_path)

    # "Neuro" only appears on the excluded prior-year week, so it must never
    # have been imported as a rotation at all.
    assert session.query(Rotation).filter_by(name="Neuro").one_or_none() is None
    assert session.query(Block).filter_by(start_date=_PRIOR_YEAR_WEEK).one_or_none() is None


def test_sync_creates_time_off_for_vacation_weeks(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [("Chen, Alice", "PGY1", [None, "VA GM", "VAC", "VAC", "AMB Endo"])],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen"])

    summary = sync(session, master_schedule_path=schedule_path, roster_path=roster_path)

    assert summary["time_off_added"] == 1
    alice = session.query(Resident).filter_by(name="Chen, Alice").one()
    time_off = session.query(TimeOff).filter_by(resident_id=alice.id).one()
    assert time_off.start_date == _WEEK_2
    assert time_off.end_date == _WEEK_3 + dt.timedelta(days=6)
    assert time_off.type == "vacation"
    assert time_off.approved is True

    # VAC weeks must never become a rotation or an assignment.
    assert session.query(Rotation).filter_by(name="VAC").one_or_none() is None
    assert session.query(Assignment).count() == 2  # only the 2 real-rotation weeks


def test_sync_is_idempotent_on_rerun(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [("Chen, Alice", "PGY1", [None, "VA GM", "VA GM", "AMB Endo", "AMB Endo"])],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen"])

    first = sync(session, master_schedule_path=schedule_path, roster_path=roster_path)
    second = sync(session, master_schedule_path=schedule_path, roster_path=roster_path)

    assert first["assignments_added"] == 4
    assert second["assignments_added"] == 0
    assert second["assignments_updated"] == 0
    assert second["time_off_added"] == 0
    assert session.query(Assignment).count() == 4
    assert session.query(Resident).count() == 1


def test_sync_skips_resident_with_no_resolvable_pgy(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [
            ("Chen, Alice", "PGY1", [None, "VA GM", "VA GM", "VA GM", "VA GM"]),
            ("Mystery, Resident", "Unknown", [None, "VA GM", "VA GM", "VA GM", "VA GM"]),
        ],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen", "Resident Mystery"])

    summary = sync(session, master_schedule_path=schedule_path, roster_path=roster_path)

    assert summary["residents"] == 1
    assert session.query(Resident).filter_by(name="Mystery, Resident").one_or_none() is None


def test_sync_raises_when_no_residents_have_pgy(tmp_path, session):
    schedule_path = _build_master_schedule_fixture(
        tmp_path,
        [("Chen, Alice", "Unknown", [None, "VA GM", "VA GM", "VA GM", "VA GM"])],
    )
    roster_path = _build_roster_fixture(tmp_path, ["Alice Chen"])

    with pytest.raises(RuntimeError):
        sync(session, master_schedule_path=schedule_path, roster_path=roster_path)
