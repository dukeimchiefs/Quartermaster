"""Tests for real_schedule/assist_list.py's readers, using small synthetic
.xlsx fixtures (fictional names) built with openpyxl.Workbook() — the one
place write-mode openpyxl is acceptable, per CLAUDE.md's real_schedule/
policy (test-fixture generation, not the production reader path).

These pin down three real bugs found live against Resident_Schedules/:
- the roster reader's end-of-table detection missed an alternate call-out
  log header wording ("Date Out" vs "Date"), so it silently ingested the
  call-out log's own data rows as garbage roster entries;
- the roster grid's own column order varies between weekly sheets ("Last
  Name, First Name, Year, Rotation" vs "Year, Last Name, First Name,
  Rotation") and was previously assumed fixed;
- load_weekly_callout_log's header matching was a single fixed-position
  tuple and missed a real alternate header layout entirely.
"""

from __future__ import annotations

import datetime as dt

import openpyxl

from real_schedule.assist_list import load_weekly_assist_roster, load_weekly_callout_log

_ROSTER_HEADER_STANDARD = (
    "Last Name",
    "First Name",
    "Year",
    "Rotation",
    "Pulls 2023-2024",
    "Pulls 2024-2025",
    "Pulls 2025-2026",
    "Pulls This Year",
    "Pulls This Block",
    "Next Rotation",
)
_ROSTER_HEADER_YEAR_FIRST = (
    "Year",
    "Last Name",
    "First Name",
    "Rotation",
    "Pulls 2022-2023",
    "Pulls 2023-2024",
    "Pulls 2024-2025",
    "Pulls This Year",
    "Pulls This Block",
    "Next Rotation",
)

_CALLOUT_HEADER_A = ("Date", "Rotation", "Resident out", "Reason", "Coverage (if assigned)")
_CALLOUT_HEADER_B = ("Date Out", "Reason Out", "Person Out", "Covering Person", "Rotation", "Notes")


def _write_week_dates(ws) -> None:
    monday = dt.date(2026, 7, 6)
    col = 11
    for day_offset in range(7):
        date_ = monday + dt.timedelta(days=day_offset)
        ws.cell(row=2, column=col, value=dt.datetime(date_.year, date_.month, date_.day))
        ws.cell(row=2, column=col + 1, value=dt.datetime(date_.year, date_.month, date_.day))
        col += 2


def _write_table(ws, start_row: int, header: tuple, rows: list[tuple]) -> int:
    for i, label in enumerate(header, start=1):
        ws.cell(row=start_row, column=i, value=label)
    row_num = start_row + 1
    for row in rows:
        for i, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=i, value=value)
        row_num += 1
    return row_num


def _build_fixture(tmp_path, sheet_name, *, roster_header, roster_rows, callout_header=None, callout_rows=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    _write_week_dates(ws)
    next_row = _write_table(ws, 4, roster_header, roster_rows)
    if callout_header is not None:
        next_row += 1  # a blank spacer row between the two tables, as in real sheets
        _write_table(ws, next_row, callout_header, callout_rows or [])
    path = tmp_path / "weekly_ASSIST_List.xlsx"
    wb.save(path)
    return str(path)


def test_load_weekly_assist_roster_standard_column_order(tmp_path):
    path = _build_fixture(
        tmp_path,
        "B1. 7.6-7.12",
        roster_header=_ROSTER_HEADER_STANDARD,
        roster_rows=[("Chen", "Alice", "PGY1", "VA GM", 0, 0, 0, 2, 0, "AMB Endo")],
    )
    records, warnings = load_weekly_assist_roster(path, "B1. 7.6-7.12")
    assert warnings == []
    assert len(records) == 1
    r = records[0]
    assert r.resident_name == "Chen, Alice"
    assert r.pgy == 1
    assert r.rotation == "VA GM"
    assert r.pulls_this_year == 2.0
    assert r.week_start == dt.date(2026, 7, 6)


def test_load_weekly_assist_roster_year_first_column_order(tmp_path):
    """Confirmed live: at least two real weekly sheets ("6.22-6.28",
    "6.15-6.21 (lauren)") list Year before Last/First Name."""
    path = _build_fixture(
        tmp_path,
        "6.22-6.28",
        roster_header=_ROSTER_HEADER_YEAR_FIRST,
        roster_rows=[("PGY-3", "Adedipe", "Oyinkansola", "CS PRT", 0, 4, 5, 3, 0, "VAC")],
    )
    records, warnings = load_weekly_assist_roster(path, "6.22-6.28")
    assert warnings == []
    assert len(records) == 1
    r = records[0]
    assert r.resident_name == "Adedipe, Oyinkansola"
    assert r.pgy == 3
    assert r.rotation == "CS PRT"
    assert r.pulls_this_year == 3.0


def test_load_weekly_assist_roster_stops_before_callout_log_alternate_header(tmp_path):
    """The exact bug found live: a call-out log header reading "Date Out"
    (not "Date") must still be recognized as the end of the roster table,
    or its data rows get ingested as garbage roster entries."""
    path = _build_fixture(
        tmp_path,
        "B1. 7.20-7.24",
        roster_header=_ROSTER_HEADER_STANDARD,
        roster_rows=[
            ("Chen", "Alice", "PGY1", "VA GM", 0, 0, 0, 2, 0, "AMB Endo"),
            ("Diaz", "Ben", "PGY2", "JEOPARDY - I", 0, 0, 0, 1, 0, "CS Heme"),
        ],
        callout_header=_CALLOUT_HEADER_B,
        callout_rows=[
            (dt.datetime(2026, 7, 20), "Scheduling error, post-nights", "Meredith Srour", "Eli Abernathy", "DUH CICU days", "confirmed"),
        ],
    )
    records, warnings = load_weekly_assist_roster(path, "B1. 7.20-7.24")
    assert warnings == []
    assert [r.resident_name for r in records] == ["Chen, Alice", "Diaz, Ben"]


def test_load_weekly_callout_log_header_format_a(tmp_path):
    path = _build_fixture(
        tmp_path,
        "B1. 7.6-7.12",
        roster_header=_ROSTER_HEADER_STANDARD,
        roster_rows=[("Chen", "Alice", "PGY1", "VA GM", 0, 0, 0, 2, 0, "AMB Endo")],
        callout_header=_CALLOUT_HEADER_A,
        callout_rows=[
            (dt.datetime(2026, 7, 8), "Duke NF", "Diaz, Ben", "Funeral", "Chen, Alice"),
        ],
    )
    records, warnings = load_weekly_callout_log(path, "B1. 7.6-7.12")
    assert warnings == []
    assert len(records) == 1
    entry = records[0]
    assert entry.date == dt.date(2026, 7, 8)
    assert entry.rotation == "Duke NF"
    assert entry.resident_out == "Diaz, Ben"
    assert entry.reason == "Funeral"
    assert entry.coverage == "Chen, Alice"


def test_load_weekly_callout_log_header_format_b(tmp_path):
    """Confirmed live: at least one real weekly sheet uses this entirely
    different header wording and column order for the same table."""
    path = _build_fixture(
        tmp_path,
        "B1. 7.20-7.24",
        roster_header=_ROSTER_HEADER_STANDARD,
        roster_rows=[("Chen", "Alice", "PGY1", "VA GM", 0, 0, 0, 2, 0, "AMB Endo")],
        callout_header=_CALLOUT_HEADER_B,
        callout_rows=[
            (dt.datetime(2026, 7, 20), "Scheduling error, post-nights", "Diaz, Ben", "Chen, Alice", "DUH CICU days", "confirmed"),
        ],
    )
    records, warnings = load_weekly_callout_log(path, "B1. 7.20-7.24")
    assert warnings == []
    assert len(records) == 1
    entry = records[0]
    assert entry.date == dt.date(2026, 7, 20)
    assert entry.rotation == "DUH CICU days"
    assert entry.resident_out == "Diaz, Ben"
    assert entry.reason == "Scheduling error, post-nights"
    assert entry.coverage == "Chen, Alice"


def test_load_weekly_callout_log_none_found_is_a_warning_not_a_crash(tmp_path):
    path = _build_fixture(
        tmp_path,
        "6.22-6.28",
        roster_header=_ROSTER_HEADER_YEAR_FIRST,
        roster_rows=[("PGY-3", "Adedipe", "Oyinkansola", "CS PRT", 0, 4, 5, 3, 0, "VAC")],
    )
    records, warnings = load_weekly_callout_log(path, "6.22-6.28")
    assert records == []
    assert len(warnings) == 1
    assert "not found" in warnings[0].reason
