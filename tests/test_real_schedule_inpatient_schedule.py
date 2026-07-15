"""Tests for real_schedule/inpatient_schedule.py's reader, using small
synthetic .xlsx fixtures (fictional names) built with openpyxl.Workbook()
— the one place write-mode openpyxl is acceptable, per CLAUDE.md's
real_schedule/ policy (test-fixture generation, not the production reader
path).

Pin down the real structural variance confirmed live (2026-07) across 5
inpatient service files: day-name header text is optional (at least one
real file has none at all — the date row alone is what matters), leading
label-column count varies (2 vs 3 columns before the resident name), and
at least one real file has a second, stale prior-year date section
sitting in later columns of the same header row.
"""

from __future__ import annotations

import datetime as dt

import openpyxl

from real_schedule.inpatient_schedule import load_inpatient_week_rows
from real_schedule.roster import RosterEntry, RosterIndex

_MONDAY = dt.datetime(2026, 7, 6)


def _write_block(ws, start_row: int, *, name_col: int, team_col: int, rows: list[tuple], day_header_text: list[str] | None = None):
    """name_col/team_col are 1-indexed. Day columns are name_col+1 .. name_col+7."""
    header_row = start_row
    if day_header_text is not None:
        for i, label in enumerate(day_header_text):
            ws.cell(row=header_row, column=name_col + 1 + i, value=label)
        header_row += 1
    for i in range(7):
        date_ = _MONDAY + dt.timedelta(days=i)
        ws.cell(row=header_row, column=name_col + 1 + i, value=date_)
    row_num = header_row + 1
    for team, name, *shifts in rows:
        ws.cell(row=row_num, column=team_col, value=team)
        ws.cell(row=row_num, column=name_col, value=name)
        for i, shift in enumerate(shifts):
            ws.cell(row=row_num, column=name_col + 1 + i, value=shift)
        row_num += 1
    return row_num


def test_load_inpatient_week_rows_with_day_name_header(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_block(
        ws, 1, name_col=3, team_col=2,
        day_header_text=["Mon", "Tues", "Wed", "Thurs", "Fri", "Sat", "Sun"],
        rows=[("GM1", "Chen, Alice", "On", "Post", "Short", "Pre", "On", "OFF", "Post")],
    )
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    records, warnings = load_inpatient_week_rows(str(path), "Sheet")
    assert warnings == []
    assert len(records) == 1
    r = records[0]
    assert r.team == "GM1"
    assert r.resident_name == "Chen, Alice"
    assert r.day_parts[dt.date(2026, 7, 11)] == "OFF"  # Saturday


def test_load_inpatient_week_rows_without_day_name_header(tmp_path):
    """Confirmed live: at least one real file (DUH CICU) has no day-name
    text row at all — the date row itself is what locates the block."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_block(
        ws, 1, name_col=3, team_col=2,
        day_header_text=None,
        rows=[("INTERN", "Diaz, Ben", "Day A", "Day A", "Day A", "OFF", "Day A", "Day A", "Day A")],
    )
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    records, warnings = load_inpatient_week_rows(str(path), "Sheet")
    assert warnings == []
    assert len(records) == 1
    assert records[0].day_parts[dt.date(2026, 7, 9)] == "OFF"  # Thursday


def test_load_inpatient_week_rows_handles_different_leading_column_counts(tmp_path):
    """Confirmed live: some files have 2 leading label columns (team,
    name), others 3 (attending, team, name) — the name/team column
    position is derived from the date columns, never fixed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_block(
        ws, 1, name_col=4, team_col=3,  # an extra "attending" column at 1-2
        day_header_text=None,
        rows=[("GM1", "Botros, Fady", "On", "Post", "Short", "Pre", "On", "OFF", "Post")],
    )
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    records, warnings = load_inpatient_week_rows(str(path), "Sheet")
    assert len(records) == 1
    assert records[0].resident_name == "Botros, Fady"
    assert records[0].team == "GM1"


def test_load_inpatient_week_rows_skips_spacer_and_section_label_rows(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    end_row = _write_block(
        ws, 1, name_col=3, team_col=2,
        day_header_text=None,
        rows=[("GM1", "Chen, Alice", "On", "Post", "Short", "Pre", "On", "OFF", "Post")],
    )
    # A section-label row: text in the name column, but no real shift data
    # in the day columns (matching real "HOUSESTAFF TEAMS"-style rows).
    ws.cell(row=end_row, column=3, value="SECTION LABEL")
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    records, warnings = load_inpatient_week_rows(str(path), "Sheet")
    assert len(records) == 1
    assert records[0].resident_name == "Chen, Alice"


def test_load_inpatient_week_rows_canonicalizes_against_roster(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _write_block(
        ws, 1, name_col=3, team_col=2,
        day_header_text=None,
        rows=[("GM1", "Alice Chen", "On", "Post", "Short", "Pre", "On", "OFF", "Post")],
    )
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    roster_index = RosterIndex([RosterEntry(canonical_name="Chen, Alice", first="Alice", last="Chen")])
    records, warnings = load_inpatient_week_rows(str(path), "Sheet", roster=roster_index)
    assert warnings == []
    assert records[0].resident_name == "Chen, Alice"


def test_load_inpatient_week_rows_two_week_blocks(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    end_row = _write_block(
        ws, 1, name_col=3, team_col=2,
        day_header_text=None,
        rows=[("GM1", "Chen, Alice", "On", "Post", "Short", "Pre", "On", "OFF", "Post")],
    )
    ws.cell(row=end_row + 1, column=1, value="BLOCK 2")
    second_monday = _MONDAY + dt.timedelta(days=7)
    for i in range(7):
        ws.cell(row=end_row + 2, column=4 + i, value=second_monday + dt.timedelta(days=i))
    ws.cell(row=end_row + 3, column=2, value="GM1")
    ws.cell(row=end_row + 3, column=3, value="Chen, Alice")
    for i, shift in enumerate(["Post", "Short", "Pre", "On", "OFF", "Post", "Short"]):
        ws.cell(row=end_row + 3, column=4 + i, value=shift)
    path = tmp_path / "inpatient.xlsx"
    wb.save(path)

    records, warnings = load_inpatient_week_rows(str(path), "Sheet")
    assert len(records) == 2
    week_starts = sorted({min(r.day_parts) for r in records})
    assert week_starts == [dt.date(2026, 7, 6), dt.date(2026, 7, 13)]


def test_load_inpatient_week_rows_sheet_not_found(tmp_path):
    wb = openpyxl.Workbook()
    path = tmp_path / "no_such_sheet.xlsx"
    wb.save(path)
    records, warnings = load_inpatient_week_rows(str(path), "Nonexistent")
    assert records == []
    assert len(warnings) == 1
    assert "not found" in warnings[0].reason
