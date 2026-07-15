"""Tests for real_schedule/fsc_tracker.py's reader, using a small synthetic
.xlsx fixture (fictional names) built with openpyxl.Workbook() — the one
place write-mode openpyxl is acceptable, per CLAUDE.md's real_schedule/
policy (test-fixture generation, not the production reader path).
"""

from __future__ import annotations

import openpyxl

from real_schedule.fsc_tracker import load_fsc_tracker
from real_schedule.roster import RosterEntry, RosterIndex

_HEADER = (
    "Resident",
    "Program",
    "PGY",
    "NetID",
    "Base FSC",
    "Extra Total",
    "Extra Reason",
    "FSC Available",
    "FSC Used",
    "FSC Left",
    "Current Phase",
    "Column1",
)


def _build_fixture(tmp_path, rows: list[tuple]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FSCTracker"
    for i, label in enumerate(_HEADER, start=1):
        ws.cell(row=1, column=i, value=label)
    for row_num, row in enumerate(rows, start=2):
        for i, value in enumerate(row, start=1):
            ws.cell(row=row_num, column=i, value=value)
    path = tmp_path / "FSC_ConfDay_Trackers.xlsx"
    wb.save(path)
    return str(path)


def test_load_fsc_tracker_happy_path(tmp_path):
    path = _build_fixture(
        tmp_path,
        [
            ("Chen, Alice", "Categorical", "PGY-1", "abc1", 4, None, None, 4, 0, 4, "Appointment Time", None),
            ("Diaz, Ben", "Categorical", "PGY-2", "def2", 4, None, None, 4, 2.5, 1.5, "Appointment Time", None),
        ],
    )
    records, warnings = load_fsc_tracker(path)
    assert warnings == []
    assert len(records) == 2
    r = records[0]
    assert r.resident_name == "Chen, Alice"
    assert r.pgy == 1
    assert r.program == "Categorical"
    assert r.base_fsc == 4.0
    assert r.fsc_available == 4.0
    assert r.fsc_used == 0.0
    assert r.fsc_left == 4.0
    assert r.phase == "Appointment Time"
    assert records[1].fsc_left == 1.5


def test_load_fsc_tracker_skips_blank_rows(tmp_path):
    path = _build_fixture(
        tmp_path,
        [
            ("Chen, Alice", "Categorical", "PGY-1", "abc1", 4, None, None, 4, 0, 4, "Appointment Time", None),
            (None, None, None, None, None, None, None, None, None, None, None, None),
        ],
    )
    records, warnings = load_fsc_tracker(path)
    assert warnings == []
    assert len(records) == 1


def test_load_fsc_tracker_warns_not_crashes_on_bad_numeric_value(tmp_path):
    """Confirmed live elsewhere in real_schedule/ that broken Excel formula
    references surface as literal error strings (e.g. "#N/A") — this must
    degrade to a ParseWarning, not a crash, same as assist_list.py's
    Pulls-This-Year handling."""
    path = _build_fixture(
        tmp_path,
        [("Chen, Alice", "Categorical", "PGY-1", "abc1", 4, None, None, 4, "#N/A", "#N/A", "Appointment Time", None)],
    )
    records, warnings = load_fsc_tracker(path)
    assert len(records) == 1
    assert records[0].fsc_used is None
    assert records[0].fsc_left is None
    assert len(warnings) == 2


def test_load_fsc_tracker_canonicalizes_against_roster(tmp_path):
    path = _build_fixture(
        tmp_path,
        [("Alice Chen", "Categorical", "PGY-1", "abc1", 4, None, None, 4, 0, 4, "Appointment Time", None)],
    )
    roster_index = RosterIndex([RosterEntry(canonical_name="Chen, Alice", first="Alice", last="Chen")])
    records, warnings = load_fsc_tracker(path, roster=roster_index)
    assert warnings == []
    assert records[0].resident_name == "Chen, Alice"


def test_load_fsc_tracker_sheet_not_found(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "SomeOtherSheet"
    path = tmp_path / "no_fsc_sheet.xlsx"
    wb.save(path)
    records, warnings = load_fsc_tracker(str(path))
    assert records == []
    assert len(warnings) == 1
    assert "not found" in warnings[0].reason
