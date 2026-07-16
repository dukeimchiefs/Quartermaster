"""Page 4 — Check Assist Swap.

Verifies a proposed assist/jeopardy week-swap (two residents trading which
week each is on backup duty) against the real, live schedule workbooks at
Resident_Schedules/ — read-only, per CLAUDE.md's PII boundary. This page
never commits or writes anything back to those workbooks, or to db.models
(there was never a DB row to create — this data lives in Excel, not the
toy SQLite DB the rest of this app uses). A "check" here is this app's
analogue of "propose" elsewhere: every run writes to audit_log, since
CLAUDE.md requires every proposed change to be logged, not just committed
ones.

Inputs deliberately mirror the real `Assist List Swaps` sheet's own
columns (Resident #1/#2, the week being covered, resident #1's new week) —
filling out this form is the same four fields a chief already writes into
that log by hand.

Has two ways to run a check — the structured form above and a free-text box
parsed by the local assistant (llm.tools.handle_check_assist_swap_message)
— both converge on the same check_assist_swap() result and the same
render/review section below, so results are stored in st.session_state
rather than rendered directly inside either button's if-block (Streamlit
reruns the whole script on every widget interaction, and a button nested
inside that block would go stale on the very rerun meant to handle its
click).

Nested "Review": since this page is read-only over Excel by design, review
means recording an Approve/Reject decision to audit_log — the chief still
has to update the real Assist List Swaps sheet by hand.
"""

from __future__ import annotations

import datetime as dt
import json
import os

import streamlit as st

from app.auth import get_actor, require_chief_auth
from audit.log import record as audit_record
from llm.tools import handle_check_assist_swap_message
from real_schedule.assist_list import load_master_assist_list, load_weekly_assist_roster
from real_schedule.checks import check_assist_swap
from real_schedule.master_schedule import load_master_schedule
from real_schedule.roster import RosterIndex, load_roster

require_chief_auth()

st.title("Check Assist Swap")
st.caption("Verify a proposed jeopardy/assist week-swap against the real, live schedules (read-only).")

_RESIDENT_SCHEDULES_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "Resident_Schedules")
_MASTER_ASSIST_PATH = os.path.join(_RESIDENT_SCHEDULES_DIR, "master_ASSIST_List_2026-2027.xlsx")
_WEEKLY_ASSIST_PATH = os.path.join(_RESIDENT_SCHEDULES_DIR, "weekly_ASSIST_List_2026-2027.xlsx")
_MASTER_SCHEDULE_PATH = os.path.join(_RESIDENT_SCHEDULES_DIR, "master_MASTER_Schedule_2026-2027.xlsx")
_ROSTER_PATH = os.path.join(_RESIDENT_SCHEDULES_DIR, "duke_residency_2026-2027.csv")


@st.cache_data(show_spinner="Reading real schedule workbooks...")
def _load_all(master_assist_mtime: float, weekly_assist_mtime: float, master_schedule_mtime: float, roster_mtime: float):
    """mtimes are part of the cache key so an updated real workbook (the
    OneDrive sync writes a new file, it isn't edited in place) invalidates
    the cache automatically — Streamlit's cache_data can't see file
    changes on its own since load_workbook only takes a path."""
    del master_assist_mtime, weekly_assist_mtime, master_schedule_mtime, roster_mtime  # cache-key only
    roster_entries, w0 = load_roster(_ROSTER_PATH)
    roster = RosterIndex(roster_entries)
    master_assist, w1 = load_master_assist_list(_MASTER_ASSIST_PATH, roster=roster)
    master_schedule, w2 = load_master_schedule(_MASTER_SCHEDULE_PATH, roster=roster)
    import openpyxl

    wb = openpyxl.load_workbook(_WEEKLY_ASSIST_PATH, read_only=True)
    weekly_sheet_names = [
        name for name in wb.sheetnames if name not in ("How To", "Pull Counter", "Sick Counter", "Pre-assists", "Assist List Swaps", "Template")
    ]
    weekly_assist = []
    warnings = w0 + w1 + w2
    for sheet_name in weekly_sheet_names:
        records, w = load_weekly_assist_roster(_WEEKLY_ASSIST_PATH, sheet_name, roster=roster)
        weekly_assist.extend(records)
        warnings.extend(w)
    return master_assist, weekly_assist, master_schedule, warnings


if not os.path.isdir(_RESIDENT_SCHEDULES_DIR):
    st.error(f"Resident_Schedules/ not found at {_RESIDENT_SCHEDULES_DIR} — can't read the real schedules.")
    st.stop()

try:
    mtimes = tuple(
        os.path.getmtime(p) for p in (_MASTER_ASSIST_PATH, _WEEKLY_ASSIST_PATH, _MASTER_SCHEDULE_PATH, _ROSTER_PATH)
    )
except OSError as exc:
    st.error(f"Couldn't read one of the real schedule workbooks: {exc}")
    st.stop()

master_assist, weekly_assist, master_schedule, load_warnings = _load_all(*mtimes)

if load_warnings:
    with st.expander(f"{len(load_warnings)} parsing warning(s) while reading the real workbooks"):
        for w in load_warnings[:50]:
            st.caption(f"{w.sheet} (row {w.row}): {w.reason}")

resident_names = sorted({d.resident_name for d in master_assist} | {e.resident_name for e in weekly_assist})
week_starts = sorted({e.week_start for e in weekly_assist})

if not resident_names or not week_starts:
    st.warning("No residents or weeks could be parsed from the real schedule workbooks.")
    st.stop()


def _result_to_dict(result, *, inputs: dict, source: str) -> dict:
    return {
        "source": source,
        "inputs": inputs,
        "is_clear": result.is_clear,
        "findings": [{"rule": f.rule, "severity": f.severity, "message": f.message} for f in result.findings],
        "reminders": list(result.reminders),
    }


def _stage_result(result_dict: dict) -> None:
    st.session_state["assist_swap_result"] = result_dict
    st.session_state.pop("assist_swap_review_status", None)


col1, col2 = st.columns(2)
with col1:
    resident_1 = st.selectbox("Resident #1 (currently on jeopardy/assist)", options=resident_names)
with col2:
    resident_2 = st.selectbox("Resident #2 (proposed to cover)", options=resident_names, index=min(1, len(resident_names) - 1))

col3, col4 = st.columns(2)
with col3:
    week_covered = st.selectbox(
        "Week resident #2 would cover",
        options=week_starts,
        format_func=lambda d: d.strftime("%b %-d, %Y"),
    )
with col4:
    week_new = st.selectbox(
        "Resident #1's new week instead",
        options=week_starts,
        format_func=lambda d: d.strftime("%b %-d, %Y"),
        index=min(1, len(week_starts) - 1),
    )

if st.button("Check this swap", type="primary"):
    result = check_assist_swap(
        resident_1, resident_2, week_covered, week_new,
        master_assist=master_assist, weekly_assist=weekly_assist, master_schedule=master_schedule,
    )
    inputs = {
        "resident_1": resident_1, "resident_2": resident_2,
        "week_covered": week_covered.isoformat(), "week_new": week_new.isoformat(),
    }
    audit_record(
        actor=get_actor(),
        action="check_assist_swap",
        reason=f"checked proposed swap: {resident_1} <-> {resident_2}, covering {week_covered}, new week {week_new}",
        details=json.dumps({**inputs, "is_clear": result.is_clear, "findings": [
            {"rule": f.rule, "severity": f.severity, "message": f.message} for f in result.findings
        ]}),
    )
    _stage_result(_result_to_dict(result, inputs=inputs, source="structured"))

st.divider()
st.subheader("Or describe it in your own words")
st.caption(
    "Parses free text into the two residents and weeks via the local assistant, then always runs the same "
    "real_schedule.checks.check_assist_swap() as the form above — it never invents a verdict itself."
)
free_text = st.text_area(
    "What's the proposed swap?",
    placeholder='e.g. "Sarah covers Tom\'s jeopardy week of 7/13, Tom takes the week of 7/20 instead"',
    key="assist_swap_free_text",
)
if st.button("Parse & check"):
    if not free_text.strip():
        st.warning("Enter a description first.")
    else:
        try:
            with st.spinner("Asking the local assistant..."):
                handled = handle_check_assist_swap_message(
                    master_assist, weekly_assist, master_schedule, free_text, today=dt.date.today()
                )
        except Exception as exc:  # Ollama not running, model not pulled, etc.
            st.error(f"Local assistant unavailable ({exc}). Use the structured form above instead.")
        else:
            if not handled.resolved:
                st.info(handled.reply)
            else:
                result = handled.result
                inputs = {
                    "resident_1": handled.resolved_args["resident_1"], "resident_2": handled.resolved_args["resident_2"],
                    "week_covered": handled.resolved_args["week_covered"].isoformat(),
                    "week_new": handled.resolved_args["week_new"].isoformat(),
                    "free_text": free_text.strip(),
                }
                audit_record(
                    actor=get_actor(),
                    action="check_assist_swap",
                    reason=f'checked assist swap (free text): "{free_text.strip()}"',
                    details=json.dumps({**inputs, "is_clear": result.is_clear, "findings": [
                        {"rule": f.rule, "severity": f.severity, "message": f.message} for f in result.findings
                    ]}),
                )
                result_dict = _result_to_dict(result, inputs=inputs, source="free_text")
                result_dict["narration"] = handled.reply
                _stage_result(result_dict)

stored = st.session_state.get("assist_swap_result")
if stored:
    st.divider()
    if stored["source"] == "free_text":
        st.success(stored["narration"])
    if stored["is_clear"]:
        st.success("No blocking issues found.")
    else:
        st.error("This swap has at least one blocking issue — review before proceeding.")

    for finding in stored["findings"]:
        container = st.error if finding["severity"] == "blocking" else st.warning
        container(finding["message"])

    if stored["reminders"]:
        st.divider()
        st.caption("Reminders (not machine-checkable):")
        for reminder in stored["reminders"]:
            st.info(reminder)

    st.divider()
    st.subheader("Review")
    review_status = st.session_state.get("assist_swap_review_status")
    if review_status:
        st.info(f"Already recorded as **{review_status}**.")
    else:
        st.caption(
            "This app never writes to Resident_Schedules/ — approving here only records the chief's decision "
            "to the audit log. Update the real Assist List Swaps sheet yourself."
        )
        rev_col1, rev_col2 = st.columns(2)
        with rev_col1:
            approve = st.button("Approve", type="primary", key="approve_assist_swap")
        with rev_col2:
            reject = st.button("Reject", key="reject_assist_swap")
        if approve:
            audit_record(
                actor=get_actor(), action="approve_assist_swap",
                reason=f"approved assist swap: {stored['inputs']}", details=json.dumps(stored),
            )
            st.session_state["assist_swap_review_status"] = "approved"
            st.rerun()
        if reject:
            audit_record(
                actor=get_actor(), action="reject_assist_swap",
                reason=f"rejected assist swap: {stored['inputs']}", details=json.dumps(stored),
            )
            st.session_state["assist_swap_review_status"] = "rejected"
            st.rerun()
