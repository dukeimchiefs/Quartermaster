"""Reader for master_AMBULATORY_schedule_2026-2027.xlsx's per-week sheets
(e.g. "B1. 7.6-7.10") — Monday-Friday only, no weekend columns (ambulatory
clinic doesn't run weekends).

Verified live (2026-07) against the real file: header row 6, data from
row 7, dates-per-day live in row 5 (one date per AM/PM column pair, at the
AM column position — the PM column for that day shares the same date).
A populated day-part cell is either a subspecialty-preceptor cell
("Name\\n(SITE CODE)", see common.is_preceptor_cell) or a bare CC-panel/
admin placeholder for the resident's own continuity clinic — both are kept
as raw text here; disambiguating them is checks.py's job, not this reader's.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

import openpyxl

from real_schedule.common import ParseWarning, normalize_pgy

_HEADER_ROW = 6
_DATA_START_ROW = 7
_DATE_ROW = 5
_DAY_PART_COLUMN_NAMES = (
    "Mon AM",
    "Mon PM",
    "Tues AM",
    "Tues PM",
    "Wed AM",
    "Wed PM",
    "Thurs AM",
    "Thurs PM",
    "Fri AM",
    "Fri PM",
)


@dataclass(frozen=True)
class AmbulatoryWeekRow:
    resident_name: str
    pgy: int | None
    rotation: str | None
    day_parts: dict[tuple[dt.date, str], str] = field(default_factory=dict)


def _column_index_map(header_row: tuple) -> dict[str, int]:
    """1-indexed column number for each named header, first occurrence
    only (the sheet reuses "CC" as both a day-part-adjacent column and a
    trailing summary column — only the first, day-part-relevant block of
    headers matters here)."""
    index_map: dict[str, int] = {}
    for col_index, value in enumerate(header_row, start=1):
        if value is None:
            continue
        name = str(value).strip()
        if name not in index_map:
            index_map[name] = col_index
    return index_map


def load_ambulatory_week(path: str, sheet_name: str) -> tuple[list[AmbulatoryWeekRow], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    warnings: list[ParseWarning] = []
    if sheet_name not in wb.sheetnames:
        return [], [ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook")]
    ws = wb[sheet_name]

    header_row = next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))
    columns = _column_index_map(header_row)
    date_row = next(ws.iter_rows(min_row=_DATE_ROW, max_row=_DATE_ROW, values_only=True))

    # Resolve each day-part column name to (column_index, date) — the date
    # lives in the AM column's position in row 5; the PM column shares it.
    day_part_cols: list[tuple[str, int, dt.date]] = []
    current_date: dt.date | None = None
    for name in _DAY_PART_COLUMN_NAMES:
        col_index = columns.get(name)
        if col_index is None:
            warnings.append(ParseWarning(sheet=sheet_name, row=_HEADER_ROW, reason=f"expected column {name!r} not found"))
            continue
        if col_index - 1 < len(date_row):
            value = date_row[col_index - 1]
            if isinstance(value, (dt.datetime, dt.date)):
                current_date = value.date() if isinstance(value, dt.datetime) else value
        if current_date is None:
            warnings.append(ParseWarning(sheet=sheet_name, row=_DATE_ROW, reason=f"no date resolved for column {name!r}"))
            continue
        day_part_cols.append((name.split()[-1], col_index, current_date))  # "AM"/"PM"

    last_col_col, rotation_col = columns.get("Last Name"), columns.get("Rotation")
    first_col, year_col = columns.get("First Name"), columns.get("Year")

    records: list[AmbulatoryWeekRow] = []
    for row_number, row in enumerate(
        ws.iter_rows(min_row=_DATA_START_ROW, max_row=ws.max_row, values_only=True), start=_DATA_START_ROW
    ):
        if last_col_col is None or first_col is None:
            break
        last = row[last_col_col - 1] if last_col_col - 1 < len(row) else None
        first = row[first_col - 1] if first_col - 1 < len(row) else None
        if last is None or first is None or not str(last).strip() or not str(first).strip():
            continue  # spacer row
        name = f"{str(last).strip()}, {str(first).strip()}"
        pgy = normalize_pgy(row[year_col - 1]) if year_col and year_col - 1 < len(row) else None
        rotation_value = row[rotation_col - 1] if rotation_col and rotation_col - 1 < len(row) else None
        rotation = str(rotation_value).strip() if rotation_value is not None and str(rotation_value).strip() else None

        day_parts: dict[tuple[dt.date, str], str] = {}
        for half, col_index, date_ in day_part_cols:
            if col_index - 1 >= len(row):
                continue
            cell = row[col_index - 1]
            if cell is not None and str(cell).strip():
                day_parts[(date_, half)] = str(cell).strip()

        records.append(AmbulatoryWeekRow(resident_name=name, pgy=pgy, rotation=rotation, day_parts=day_parts))

    return records, warnings
