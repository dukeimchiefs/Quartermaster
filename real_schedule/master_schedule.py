"""Reader for master_MASTER_Schedule_2026-2027.xlsx — the year-long,
week-granular rotation source (`Intern Master` / `Upper Level Master`
sheets each have one COLUMN PER WEEK, not per 4-week block, spanning the
whole academic year) plus `Validation Lists`, the authoritative ~130-value
rotation-name taxonomy this workbook validates its own dropdowns against.

Verified live (2026-07) against the real file: header row 3, data from
row 4; week-date columns start at column L and are plain datetime values
(the Monday of that week) — no label parsing needed for this sheet, unlike
the "B1. 7.6-7.12"-style sheet names used elsewhere in Resident_Schedules/.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass

import openpyxl

from real_schedule.common import ParseWarning, normalize_pgy

_HEADER_ROW = 3
_DATA_START_ROW = 4
_NAME_COL = 1
_YEAR_COL = 5
_FIRST_WEEK_COL = 12  # column L, 1-indexed


@dataclass(frozen=True)
class MasterScheduleWeek:
    resident_name: str
    pgy: int | None
    week_start: dt.date
    rotation: str


def _read_sheet(ws, sheet_name: str) -> tuple[list[MasterScheduleWeek], list[ParseWarning]]:
    records: list[MasterScheduleWeek] = []
    warnings: list[ParseWarning] = []

    header = next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))
    week_cols: list[tuple[int, dt.date]] = []
    for col_index in range(_FIRST_WEEK_COL, len(header) + 1):
        value = header[col_index - 1]
        if isinstance(value, dt.datetime):
            week_cols.append((col_index, value.date()))
        elif isinstance(value, dt.date):
            week_cols.append((col_index, value))

    for row_number, row in enumerate(
        ws.iter_rows(min_row=_DATA_START_ROW, max_row=ws.max_row, values_only=True), start=_DATA_START_ROW
    ):
        if len(row) < _YEAR_COL:
            continue
        name = row[_NAME_COL - 1]
        if name is None or not str(name).strip():
            continue  # spacer/blank row
        name = str(name).strip()
        pgy = normalize_pgy(row[_YEAR_COL - 1])

        for col_index, week_start in week_cols:
            if col_index - 1 >= len(row):
                continue
            rotation = row[col_index - 1]
            if rotation is None or not str(rotation).strip():
                continue
            records.append(
                MasterScheduleWeek(resident_name=name, pgy=pgy, week_start=week_start, rotation=str(rotation).strip())
            )

    return records, warnings


def load_master_schedule(path: str) -> tuple[list[MasterScheduleWeek], list[ParseWarning]]:
    """Reads both Intern Master and Upper Level Master (they're identically
    shaped) and returns the combined resident x week x rotation records."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_records: list[MasterScheduleWeek] = []
    all_warnings: list[ParseWarning] = []
    for sheet_name in ("Intern Master", "Upper Level Master"):
        if sheet_name not in wb.sheetnames:
            all_warnings.append(ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook"))
            continue
        records, warnings = _read_sheet(wb[sheet_name], sheet_name)
        all_records.extend(records)
        all_warnings.extend(warnings)
    return all_records, all_warnings


def load_validation_taxonomy(path: str) -> set[str]:
    """The authoritative rotation-name taxonomy from the `Validation Lists`
    sheet's `Rotation` column (column F, header row 2) — used to sanity
    check that a rotation value scanned elsewhere is a real, known one."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Validation Lists" not in wb.sheetnames:
        return set()
    ws = wb["Validation Lists"]
    taxonomy: set[str] = set()
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=6, max_col=6, values_only=True):
        value = row[0]
        if value is not None and str(value).strip():
            taxonomy.add(str(value).strip())
    return taxonomy
