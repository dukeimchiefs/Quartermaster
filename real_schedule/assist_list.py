"""Readers for weekly_ASSIST_List_2026-2027.xlsx and
master_ASSIST_List_2026-2027.xlsx.

Verified live (2026-07) against the real files. Each weekly sheet in
weekly_ASSIST_List (e.g. "B1. 7.6-7.12") actually holds TWO separate
tables, not one:

1. The main roster grid (header row 4, data from row 5): one row per
   resident, Pulls-history/fairness columns, then a Mon-Sun AM/PM cell per
   day. `Rotation` reads JEOPARDY/JEOPARDY - I/JEOPARDY - O when that
   resident is on backup duty that week (see common.is_jeopardy_label).

2. A second, separate "call-out log" table further down the same sheet —
   this is the actual, direct record of who called out that week, for what,
   and who (if anyone) covered. This is exactly the "who is actually being
   called out" data, more direct than inferring it from the roster grid's
   day-part cells. Its row range isn't fixed (the roster grid above it is a
   different length every week), and — confirmed live — its header's
   wording AND column order both vary between weeks: one sampled week reads
   "Date, Rotation, Resident out, Reason, Coverage (if assigned)", another
   reads "Date Out, Reason Out, Person Out, Covering Person, Rotation,
   Notes". This reader locates the header by matching cell content against
   known field aliases (see _CALLOUT_FIELD_ALIASES) rather than assuming a
   fixed row number or column order. Confirmed live that this table is NOT
   always clean: one sampled week had a row that didn't match the expected
   shape at all (free text mashed into the date column) — that degrades to
   a ParseWarning, not a crash.

The `Reason` column is real, sensitive, confidential data per
`Assist List Rules`' "Keep call out reason confidential" policy — callers
(checks.py, the Streamlit pages) must never display or log it anywhere
except audit_log.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

import openpyxl

from real_schedule.common import ParseWarning, normalize_pgy, parse_duty_cell, parse_name_last_first

_ROSTER_HEADER_ROW = 4
_ROSTER_DATA_START_ROW = 5
_ROSTER_DAY_NAMES = ("Mon", "Tues", "Wed", "Thurs", "Fri", "Sat", "Sun")

_ROSTER_FIELD_ALIASES: dict[str, frozenset[str]] = {
    "last_name": frozenset({"last name"}),
    "first_name": frozenset({"first name"}),
    "year": frozenset({"year"}),
    "rotation": frozenset({"rotation"}),
    "pulls_this_year": frozenset({"pulls this year"}),
}
"""Confirmed live: the roster grid's own column order varies between weekly
sheets — most (all "B1./B2./B3."-prefixed, in-season weeks) read "Last
Name, First Name, Year, Rotation, ...", but pre-season transition sheets
("6.22-6.28", "6.15-6.21 (lauren)") read "Year, Last Name, First Name,
Rotation, ..." instead. Resolved by header content below, never by fixed
position."""


def _match_roster_header(row: tuple) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for col_index, value in enumerate(row):
        if value is None:
            continue
        text = str(value).strip().lower()
        for field_name, aliases in _ROSTER_FIELD_ALIASES.items():
            if text in aliases and field_name not in mapping:
                mapping[field_name] = col_index
    if "last_name" in mapping and "first_name" in mapping:
        return mapping
    return None

_CALLOUT_FIELD_ALIASES: dict[str, frozenset[str]] = {
    "date": frozenset({"date", "date out"}),
    "resident_out": frozenset({"resident out", "person out"}),
    "reason": frozenset({"reason", "reason out"}),
    "rotation": frozenset({"rotation"}),
    "coverage": frozenset({"coverage (if assigned)", "covering person"}),
}
"""Confirmed live: the call-out log table's header wording AND column order
both vary between weekly sheets — "Date, Rotation, Resident out, Reason,
Coverage (if assigned)" in one week vs "Date Out, Reason Out, Person Out,
Covering Person, Rotation, Notes" in another. Matched by header-cell content
below, never by fixed position."""


def _match_callout_header(row: tuple) -> dict[str, int] | None:
    """Return {field_name: column_index} if `row` looks like a call-out log
    header row (column A starts with "date", and at least the date/
    resident-out fields are both found among its cells), else None."""
    if not row or not isinstance(row[0], str) or not row[0].strip().lower().startswith("date"):
        return None
    mapping: dict[str, int] = {}
    for col_index, value in enumerate(row):
        if value is None:
            continue
        text = str(value).strip().lower()
        for field_name, aliases in _CALLOUT_FIELD_ALIASES.items():
            if text in aliases and field_name not in mapping:
                mapping[field_name] = col_index
    if "date" in mapping and "resident_out" in mapping:
        return mapping
    return None


@dataclass(frozen=True)
class AssistWeekEntry:
    resident_name: str
    pgy: int | None
    rotation: str
    week_start: dt.date
    pulls_this_year: float | None
    day_parts: dict[dt.date, str] = field(default_factory=dict)


@dataclass(frozen=True)
class CalloutLogEntry:
    date: dt.date
    rotation: str | None
    resident_out: str
    reason: str | None
    coverage: str | None


@dataclass(frozen=True)
class MasterAssistDuty:
    resident_name: str
    pgy_tier: str  # verbatim section label, e.g. "PGY-1", "PGY-3 +"
    week_start: dt.date
    duty: str          # primary duty tag; "" for a bare name with no annotation
    extra: str | None  # trailing text/second tag, e.g. "jeopardy-I" in "(Pickett) (jeopardy-I)"


def load_weekly_assist_roster(path: str, sheet_name: str) -> tuple[list[AssistWeekEntry], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return [], [ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook")]
    ws = wb[sheet_name]
    warnings: list[ParseWarning] = []

    header_row = next(ws.iter_rows(min_row=_ROSTER_HEADER_ROW, max_row=_ROSTER_HEADER_ROW, values_only=True))
    header_map = _match_roster_header(header_row)
    if header_map is None:
        warnings.append(
            ParseWarning(
                sheet=sheet_name,
                row=_ROSTER_HEADER_ROW,
                reason="roster header row missing recognizable Last Name/First Name columns — can't parse this sheet",
            )
        )
        return [], warnings
    last_col, first_col = header_map["last_name"], header_map["first_name"]
    year_col, rotation_col = header_map.get("year"), header_map.get("rotation")
    pulls_col = header_map.get("pulls_this_year")

    date_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    day_cols: list[tuple[int, dt.date]] = []
    current_date: dt.date | None = None
    for col_index, value in enumerate(date_row, start=1):
        if isinstance(value, (dt.datetime, dt.date)):
            current_date = value.date() if isinstance(value, dt.datetime) else value
            day_cols.append((col_index, current_date))
        elif current_date is not None and col_index <= 10 + 2 * len(_ROSTER_DAY_NAMES):
            # the PM column immediately after an AM date column shares it
            if day_cols and day_cols[-1][0] == col_index - 1:
                day_cols.append((col_index, current_date))

    if not day_cols:
        warnings.append(ParseWarning(sheet=sheet_name, row=2, reason="no weekday dates found in row 2 — can't resolve week_start"))
        return [], warnings
    week_start = day_cols[0][1]  # first (Monday AM) column's date

    records: list[AssistWeekEntry] = []
    row_number = _ROSTER_DATA_START_ROW - 1
    for row in ws.iter_rows(min_row=_ROSTER_DATA_START_ROW, max_row=ws.max_row, values_only=True):
        row_number += 1
        if row[0] is not None and isinstance(row[0], str) and row[0].strip().lower().startswith("date"):
            # The call-out log table's own header row also happens to
            # parse_name_last_first() as a plausible-looking (bogus) name
            # ("Date, Rotation") since neither half is blank — this check
            # must run before that parse, not only when it fails, or the
            # whole call-out log gets silently ingested as garbage roster
            # rows (confirmed live: this exact miss produced an
            # "unparseable Pulls This Year value 'Bre Updated on Medhub'"
            # warning before this fix). Matched by a "date"-prefix, not an
            # exact "Date" string — confirmed live that the call-out log's
            # header wording itself varies between weekly sheets ("Date" in
            # some, "Date Out" in others).
            break
        if isinstance(row[0], (dt.datetime, dt.date)):
            # Defense in depth: column A holds an actual date value only in
            # the call-out log table's data rows (below its header) — a
            # resident's last name is never a date. Confirmed live this
            # matters: one sheet's call-out log header read "Date Out" (not
            # caught by a hypothetical exact-"Date" check), and without this
            # the reader would have silently ingested that table's data rows
            # as garbage roster entries.
            break
        last = row[last_col] if last_col < len(row) else None
        first = row[first_col] if first_col < len(row) else None
        name = parse_name_last_first(last, first)
        if name is None:
            continue
        pgy = normalize_pgy(row[year_col]) if year_col is not None and year_col < len(row) else None
        rotation = (
            str(row[rotation_col]).strip() if rotation_col is not None and rotation_col < len(row) and row[rotation_col] is not None else ""
        )
        pulls_raw = row[pulls_col] if pulls_col is not None and pulls_col < len(row) else None
        try:
            pulls_this_year = float(pulls_raw) if pulls_raw is not None else None
        except (TypeError, ValueError):
            pulls_this_year = None
            warnings.append(ParseWarning(sheet=sheet_name, row=row_number, reason=f"unparseable Pulls This Year value {pulls_raw!r}"))

        day_parts: dict[dt.date, str] = {}
        for col_index, date_ in day_cols:
            if col_index - 1 >= len(row):
                continue
            cell = row[col_index - 1]
            if cell is not None and str(cell).strip():
                day_parts.setdefault(date_, str(cell).strip())

        records.append(
            AssistWeekEntry(
                resident_name=name,
                pgy=pgy,
                rotation=rotation,
                week_start=week_start,
                pulls_this_year=pulls_this_year,
                day_parts=day_parts,
            )
        )

    return records, warnings


def load_weekly_callout_log(path: str, sheet_name: str) -> tuple[list[CalloutLogEntry], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return [], [ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook")]
    ws = wb[sheet_name]
    warnings: list[ParseWarning] = []
    records: list[CalloutLogEntry] = []

    header_row_number: int | None = None
    header_map: dict[str, int] | None = None
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8, values_only=True), start=1):
        header_map = _match_callout_header(row)
        if header_map is not None:
            header_row_number = row_number
            break

    if header_row_number is None or header_map is None:
        warnings.append(ParseWarning(sheet=sheet_name, row=0, reason="call-out log table header not found"))
        return records, warnings

    max_col_needed = max(header_map.values()) + 1
    for row_number, row in enumerate(
        ws.iter_rows(min_row=header_row_number + 1, max_row=ws.max_row, max_col=max_col_needed, values_only=True),
        start=header_row_number + 1,
    ):
        if not any(row):
            break  # blank row ends the table
        date_value = row[header_map["date"]]
        resident_out = row[header_map["resident_out"]]
        rotation = row[header_map["rotation"]] if "rotation" in header_map else None
        reason = row[header_map["reason"]] if "reason" in header_map else None
        coverage = row[header_map["coverage"]] if "coverage" in header_map else None
        if not isinstance(date_value, (dt.datetime, dt.date)) or resident_out is None or not str(resident_out).strip():
            warnings.append(
                ParseWarning(sheet=sheet_name, row=row_number, reason=f"call-out log row doesn't match expected shape: {row!r}")
            )
            continue
        records.append(
            CalloutLogEntry(
                date=date_value.date() if isinstance(date_value, dt.datetime) else date_value,
                rotation=str(rotation).strip() if rotation is not None and str(rotation).strip() else None,
                resident_out=str(resident_out).strip(),
                reason=str(reason).strip() if reason is not None and str(reason).strip() else None,
                coverage=str(coverage).strip() if coverage is not None and str(coverage).strip() else None,
            )
        )

    return records, warnings


def load_master_assist_list(path: str) -> tuple[list[MasterAssistDuty], list[ParseWarning]]:
    """`Master Assist List` sheet: one column per week, rows grouped by a
    PGY-tier section label in column A ("PGY-1", "PGY-2", "PGY-3 +"), each
    subsequent row a free-text "Last, First (DUTY)" cell per week-column
    until the next section label or a blank row."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    warnings: list[ParseWarning] = []
    if "Master Assist List" not in wb.sheetnames:
        return [], [ParseWarning(sheet="Master Assist List", row=0, reason="sheet not found in workbook")]
    ws = wb["Master Assist List"]

    date_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    week_cols = [
        (i + 1, v.date() if isinstance(v, dt.datetime) else v)
        for i, v in enumerate(date_row)
        if isinstance(v, (dt.datetime, dt.date))
    ]

    records: list[MasterAssistDuty] = []
    current_tier: str | None = None
    for row_number, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        label = row[0] if row else None
        if label is not None and str(label).strip():
            current_tier = str(label).strip()
            continue
        if current_tier is None:
            continue
        for col_index, week_start in week_cols:
            if col_index - 1 >= len(row):
                continue
            cell = row[col_index - 1]
            if cell is None or not str(cell).strip():
                continue
            parsed = parse_duty_cell(cell)
            if parsed is None:
                warnings.append(ParseWarning(sheet="Master Assist List", row=row_number, reason=f"unparseable duty cell {cell!r}"))
                continue
            name, duty, extra = parsed
            records.append(
                MasterAssistDuty(resident_name=name, pgy_tier=current_tier, week_start=week_start, duty=duty, extra=extra)
            )

    return records, warnings


def load_pull_counter(path: str) -> tuple[dict[str, dict[dt.date, float]], list[ParseWarning]]:
    """`Pull Counter` sheet: resident -> {week_start: pulls_that_week}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    warnings: list[ParseWarning] = []
    if "Pull Counter" not in wb.sheetnames:
        return {}, [ParseWarning(sheet="Pull Counter", row=0, reason="sheet not found in workbook")]
    ws = wb["Pull Counter"]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    week_cols = [
        (i + 1, v.date() if isinstance(v, dt.datetime) else v)
        for i, v in enumerate(header)
        if isinstance(v, (dt.datetime, dt.date))
    ]

    result: dict[str, dict[dt.date, float]] = {}
    for row_number, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        last, first = row[0], row[1] if len(row) > 1 else None
        name = parse_name_last_first(last, first)
        if name is None:
            continue
        weekly: dict[dt.date, float] = {}
        for col_index, week_start in week_cols:
            if col_index - 1 >= len(row):
                continue
            value = row[col_index - 1]
            if value is None:
                continue
            try:
                weekly[week_start] = float(value)
            except (TypeError, ValueError):
                warnings.append(ParseWarning(sheet="Pull Counter", row=row_number, reason=f"unparseable pull count {value!r}"))
        result[name] = weekly

    return result, warnings
