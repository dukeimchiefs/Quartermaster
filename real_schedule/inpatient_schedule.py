"""Reader for weekly_INPATIENT_Schedules/*/*.xlsx — real, per-service
inpatient team rosters with day-level shift-type labels (used by the
day-off alignment checker, Tool 5).

Verified live (2026-07) against 5 representative real files across
DRH/VA/DUH (VA GM, DRH GM, VA MICU, DUH MICU, DUH CICU): a repeating
week-block structure, located by scanning for a run of >= 5 consecutive
date-typed cells in a row — never a fixed row/column position or day-name
header text, since at least one real file (DUH CICU) has no day-name text
header at all, just the date row directly. The resident-name column is
whatever's immediately left of the first date column, and the team/role
label column is immediately left of that — this generalizes across
confirmed real leading-column-count variance (2-3 columns) without
per-file special-casing. Resident-name format also varies per file
("First Last" in VA GM/VA MICU, "Last, First" in DRH GM/DUH MICU/
DUH CICU) — canonicalized via the shared roster.RosterIndex, same as every
other reader in this package.

A data row must have at least one non-blank day-part cell to be accepted
— this is what filters out section-label/group-header rows mixed into the
data range (e.g. "HOUSESTAFF TEAMS", "GM 1-4 (Cap 16, hosp attg)",
confirmed live in VA GM) without needing a bespoke header-text denylist:
those rows' day cells are either blank or a lone non-breaking-space
placeholder ("\xa0"), which str.strip() already treats as blank.
"""

from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field

import openpyxl

from real_schedule.common import ParseWarning
from real_schedule.roster import RosterIndex

_MIN_DATE_RUN = 5  # a full week is 7, but a holiday-shortened header block might have fewer


@dataclass(frozen=True)
class InpatientDayRow:
    team: str
    resident_name: str
    day_parts: dict[dt.date, str] = field(default_factory=dict)


def _find_week_blocks(ws) -> list[tuple[int, list[tuple[int, dt.date]]]]:
    """Scans every row for a run of >= _MIN_DATE_RUN date-typed cells that
    look like a single week (increasing, spanning <= 7 days) — the one
    place "is this a new week-block header" gets decided, never a fixed
    row number.

    Date cells are first split into CONTIGUOUS COLUMN RUNS before
    validation, not pooled across the whole row — confirmed live that at
    least one real file (DRH GM) has a second, stale prior-year date
    section sitting in later columns of the same header row (an old
    leftover "Int" sub-table a full year out of date); pooling all date
    cells together would span ~365 days and make the row's real, current
    week-block invisible. A stale run that itself still looks like a valid
    week (contiguous, increasing, <=7 days) is harmless even if accepted —
    its dates never match any current query."""
    blocks: list[tuple[int, list[tuple[int, dt.date]]]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        date_cells = [
            (i + 1, v.date() if isinstance(v, dt.datetime) else v)
            for i, v in enumerate(row)
            if isinstance(v, (dt.datetime, dt.date))
        ]
        if not date_cells:
            continue
        runs: list[list[tuple[int, dt.date]]] = [[date_cells[0]]]
        for prev, curr in zip(date_cells, date_cells[1:]):
            if curr[0] == prev[0] + 1:
                runs[-1].append(curr)
            else:
                runs.append([curr])
        for run in runs:
            if len(run) < _MIN_DATE_RUN:
                continue
            dates_sorted = sorted(d for _, d in run)
            if dates_sorted[-1] - dates_sorted[0] > dt.timedelta(days=7):
                continue
            if any(dates_sorted[i] >= dates_sorted[i + 1] for i in range(len(dates_sorted) - 1)):
                continue
            blocks.append((row_number, run))
    return blocks


def load_inpatient_week_rows(path: str, sheet_name: str, *, roster: RosterIndex | None = None) -> tuple[list[InpatientDayRow], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    warnings: list[ParseWarning] = []
    if sheet_name not in wb.sheetnames:
        return [], [ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook")]
    ws = wb[sheet_name]

    blocks = _find_week_blocks(ws)
    if not blocks:
        warnings.append(ParseWarning(sheet=sheet_name, row=0, reason="no week-block date header found"))
        return [], warnings

    records: list[InpatientDayRow] = []
    for block_index, (header_row, date_cols) in enumerate(blocks):
        name_col = date_cols[0][0] - 1
        team_col = name_col - 1
        # Two runs can share the same header row (confirmed live: a stale
        # prior-year date section alongside the current one) — bound by
        # the next DIFFERENT row number, not just the next list entry, or
        # a same-row pair would wrongly zero out each other's data range.
        next_different_row = next((r for r, _ in blocks[block_index + 1 :] if r > header_row), None)
        end_row = next_different_row - 1 if next_different_row is not None else ws.max_row
        if end_row < header_row + 1:
            continue

        for row_number, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=end_row, values_only=True), start=header_row + 1
        ):
            if name_col - 1 >= len(row):
                continue
            name = row[name_col - 1]
            if name is None or not str(name).strip():
                continue
            name = str(name).strip()

            day_parts: dict[dt.date, str] = {}
            for col_index, date_ in date_cols:
                if col_index - 1 >= len(row):
                    continue
                cell = row[col_index - 1]
                if cell is not None and str(cell).strip():
                    day_parts[date_] = str(cell).strip()

            if not day_parts:
                continue  # section-label/group-header row, not a real resident row

            if roster is not None:
                name, matched = roster.canonicalize(name)
                if not matched:
                    warnings.append(ParseWarning(sheet=sheet_name, row=row_number, reason=f"{name!r} not found in internal roster — using name as parsed"))

            team_value = row[team_col - 1] if team_col >= 1 and team_col - 1 < len(row) else None
            team = str(team_value).strip() if team_value is not None else ""

            records.append(InpatientDayRow(team=team, resident_name=name, day_parts=day_parts))

    return records, warnings
