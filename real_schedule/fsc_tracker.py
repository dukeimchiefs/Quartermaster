"""Reader for FSC_ConfDay_Trackers_2026-2027.xlsx's `FSCTracker` sheet — the
resident FSC ("Flexible Scholarly/Conference") day bank: how many days each
resident has, has used, and has left.

Verified live (2026-07): header row 1, data from row 2. `Resident` is the
same "Last, First" key (sometimes with a "(C1)"-style panel-number suffix)
used throughout real_schedule/ — confirmed identical strings for the same
resident against master_MASTER_Schedule. `FSC Left` takes half-integer
values (0, 1, 1.5, 2, 3, 3.5, 4, 5), confirming this bank is already tracked
at half-day granularity. Columns are resolved by header text, not fixed
position, matching ambulatory.py's approach (assist_list.py's roster reader
had a real, confirmed bug this session from assuming a fixed column order
that turned out to vary between sheets).

The workbook's other sheet, `Conference Days`, is a blank "track manually if
you want" placeholder with no real data — confirmed live, not read here.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl

from real_schedule.common import ParseWarning, normalize_pgy
from real_schedule.roster import RosterIndex

_HEADER_ROW = 1
_DATA_START_ROW = 2

_FIELD_ALIASES: dict[str, frozenset[str]] = {
    "resident": frozenset({"resident"}),
    "program": frozenset({"program"}),
    "pgy": frozenset({"pgy"}),
    "base_fsc": frozenset({"base fsc"}),
    "fsc_available": frozenset({"fsc available"}),
    "fsc_used": frozenset({"fsc used"}),
    "fsc_left": frozenset({"fsc left"}),
    "phase": frozenset({"current phase"}),
}


@dataclass(frozen=True)
class FscBalance:
    resident_name: str
    pgy: int | None
    program: str | None
    base_fsc: float | None
    fsc_available: float | None
    fsc_used: float | None
    fsc_left: float | None
    phase: str | None  # "Appointment Time" / "Reflection Time" — carried through, not filtered on


def _match_header(row: tuple) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for col_index, value in enumerate(row):
        if value is None:
            continue
        text = str(value).strip().lower()
        for field_name, aliases in _FIELD_ALIASES.items():
            if text in aliases and field_name not in mapping:
                mapping[field_name] = col_index
    if "resident" in mapping:
        return mapping
    return None


def _parse_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def load_fsc_tracker(path: str, *, roster: RosterIndex | None = None) -> tuple[list[FscBalance], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    warnings: list[ParseWarning] = []
    if "FSCTracker" not in wb.sheetnames:
        return [], [ParseWarning(sheet="FSCTracker", row=0, reason="sheet not found in workbook")]
    ws = wb["FSCTracker"]

    header_row = next(ws.iter_rows(min_row=_HEADER_ROW, max_row=_HEADER_ROW, values_only=True))
    header_map = _match_header(header_row)
    if header_map is None:
        return [], [ParseWarning(sheet="FSCTracker", row=_HEADER_ROW, reason="header row missing a recognizable Resident column")]

    records: list[FscBalance] = []
    for row_number, row in enumerate(
        ws.iter_rows(min_row=_DATA_START_ROW, max_row=ws.max_row, values_only=True), start=_DATA_START_ROW
    ):
        resident_col = header_map["resident"]
        name = row[resident_col] if resident_col < len(row) else None
        if name is None or not str(name).strip():
            continue  # spacer/blank row
        name = str(name).strip()
        if roster is not None:
            name, matched = roster.canonicalize(name)
            if not matched:
                warnings.append(ParseWarning(sheet="FSCTracker", row=row_number, reason=f"{name!r} not found in internal roster — using name as parsed"))

        pgy_col = header_map.get("pgy")
        pgy = normalize_pgy(row[pgy_col]) if pgy_col is not None and pgy_col < len(row) else None

        program_col = header_map.get("program")
        program_value = row[program_col] if program_col is not None and program_col < len(row) else None
        program = str(program_value).strip() if program_value is not None and str(program_value).strip() else None

        phase_col = header_map.get("phase")
        phase_value = row[phase_col] if phase_col is not None and phase_col < len(row) else None
        phase = str(phase_value).strip() if phase_value is not None and str(phase_value).strip() else None

        numeric_fields: dict[str, float | None] = {}
        for field_name in ("base_fsc", "fsc_available", "fsc_used", "fsc_left"):
            col = header_map.get(field_name)
            raw = row[col] if col is not None and col < len(row) else None
            parsed = _parse_float(raw)
            if raw is not None and parsed is None:
                warnings.append(ParseWarning(sheet="FSCTracker", row=row_number, reason=f"unparseable {field_name!r} value {raw!r}"))
            numeric_fields[field_name] = parsed

        records.append(
            FscBalance(
                resident_name=name,
                pgy=pgy,
                program=program,
                base_fsc=numeric_fields["base_fsc"],
                fsc_available=numeric_fields["fsc_available"],
                fsc_used=numeric_fields["fsc_used"],
                fsc_left=numeric_fields["fsc_left"],
                phase=phase,
            )
        )

    return records, warnings
