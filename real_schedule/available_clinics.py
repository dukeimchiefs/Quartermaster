"""Reader for the "available clinics" workbook — a seasonal file (e.g.
`Available Clinics Fall 2026.xlsx`; a Spring version will presumably
follow) with the `Intern Blocks` / `JAR Blocks` / `SAR-Elective Blocks`
sheets Tool 2 needs. Path is always a parameter here, never hardcoded,
since the filename changes by season.

Verified live (2026-07): header row 1, data from row 2, with section
headers (specialty group labels like "ID\\n13 Clinics / Confirmed in
Epic") interspersed as spacer rows — these have no comma in column A and
are skipped, not treated as data (the `Specialty` column on each real data
row already carries this information directly, so the section header is
purely a human-readability divider, not the only source of the specialty
value). A populated day-part cell uses the exact same "Name\\n(SITE CODE)"
shape as master_AMBULATORY_schedule's per-week sheets — common.is_preceptor_cell
is reused directly rather than inventing a second parsing convention.
"""

from __future__ import annotations

from dataclasses import dataclass, field

import openpyxl

from real_schedule.common import ParseWarning, is_preceptor_cell

_DAY_PART_COLUMN_NAMES = (
    "Mon AM",
    "Mon PM",
    "Tues AM",
    "Tues PM",
    "Wed AM",
    "Wed PM",
    "Thurs AM",
    "Thurs PM",
    "Fri AM",
    "Fri PM",
)

_TIER_SHEETS = ("Intern Blocks", "JAR Blocks", "SAR-Elective Blocks")


@dataclass(frozen=True)
class ClinicSlot:
    preceptor_name: str
    specialty: str | None
    location: str | None  # the sheet's own summary Location column (may list multiple day-specific codes)
    tier: str  # which sheet this came from: "Intern Blocks" / "JAR Blocks" / "SAR-Elective Blocks"
    available_day_parts: set[tuple[str, str]] = field(default_factory=set)  # ("Mon", "AM"), keyed by parsed site code below
    site_codes_by_day_part: dict[tuple[str, str], str] = field(default_factory=dict)
    notes: str | None = None


def _column_index_map(header_row: tuple) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for col_index, value in enumerate(header_row, start=1):
        if value is None:
            continue
        name = str(value).strip()
        if name not in index_map:
            index_map[name] = col_index
    return index_map


def _read_sheet(ws, sheet_name: str) -> tuple[list[ClinicSlot], list[ParseWarning]]:
    warnings: list[ParseWarning] = []
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    columns = _column_index_map(header)

    name_col = columns.get("Name")
    specialty_col = columns.get("Specialty")
    location_col = columns.get("Location")
    notes_col = columns.get("Notes")
    if name_col is None:
        return [], [ParseWarning(sheet=sheet_name, row=1, reason="expected column 'Name' not found")]

    day_part_cols = [(name, columns[name]) for name in _DAY_PART_COLUMN_NAMES if name in columns]

    records: list[ClinicSlot] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        if name_col - 1 >= len(row):
            continue
        raw_name = row[name_col - 1]
        if raw_name is None or "," not in str(raw_name):
            continue  # section-header/spacer row, not a preceptor data row
        preceptor_name = str(raw_name).strip()

        specialty = None
        if specialty_col and specialty_col - 1 < len(row) and row[specialty_col - 1]:
            specialty = str(row[specialty_col - 1]).strip()
        location = None
        if location_col and location_col - 1 < len(row) and row[location_col - 1]:
            location = str(row[location_col - 1]).strip()
        notes = None
        if notes_col and notes_col - 1 < len(row) and row[notes_col - 1]:
            notes = str(row[notes_col - 1]).strip()

        available: set[tuple[str, str]] = set()
        site_codes: dict[tuple[str, str], str] = {}
        for col_name, col_index in day_part_cols:
            if col_index - 1 >= len(row):
                continue
            cell = row[col_index - 1]
            if cell is None:
                continue
            day, half = col_name.split()[0], col_name.split()[1]
            parsed = is_preceptor_cell(cell)
            if parsed is not None:
                available.add((day, half))
                site_codes[(day, half)] = parsed[1]
            else:
                warnings.append(
                    ParseWarning(sheet=sheet_name, row=row_number, reason=f"day-part cell not in expected 'Name\\n(SITE)' shape: {cell!r}")
                )

        records.append(
            ClinicSlot(
                preceptor_name=preceptor_name,
                specialty=specialty,
                location=location,
                tier=sheet_name,
                available_day_parts=available,
                site_codes_by_day_part=site_codes,
                notes=notes,
            )
        )

    return records, warnings


def load_available_clinics(path: str) -> tuple[list[ClinicSlot], list[ParseWarning]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_records: list[ClinicSlot] = []
    all_warnings: list[ParseWarning] = []
    for sheet_name in _TIER_SHEETS:
        if sheet_name not in wb.sheetnames:
            all_warnings.append(ParseWarning(sheet=sheet_name, row=0, reason="sheet not found in workbook"))
            continue
        records, warnings = _read_sheet(wb[sheet_name], sheet_name)
        all_records.extend(records)
        all_warnings.extend(warnings)
    return all_records, all_warnings
